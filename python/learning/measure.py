#!/usr/bin/env python

"""
Compute training and cross-validation statistics.
"""

import os
import re
import statistics
from typing import Dict, List, Tuple

import numpy as np
import openpyxl
import openpyxl.styles as st
import pandas as pd
from sklearn.inspection import permutation_importance
from sklearn.metrics import (balanced_accuracy_score, f1_score, precision_score,
                             recall_score, confusion_matrix)
from tabulate import tabulate

from utils.util import (safe_drop, filter_df,
                        rename_features, rename_feature_set)

to_ignore_missing = False

cutoff = 0.5

FAILURES_YML = os.path.abspath(os.path.split(
    os.path.abspath(__file__))[0]+'/param_failures.yml')


def preprocess_measure(path, to_filter=False, in_feature_set=[]):
    if len(in_feature_set) == 0:
        from learning.feature_set import feature_set, typical_feature_set, secondary_feature_set

    if 'csv' in path:
        in_data = pd.read_csv(path)
    else:
        in_data = pd.read_excel(path)

    is_typical = 'stroke_indices2' not in in_data.columns
    is_global = len([col for col in in_data.columns if 'global_' in col]) > 0
    if is_typical and len(in_feature_set) == 0:
        feature_set = typical_feature_set
    if is_global and len(in_feature_set) == 0:
        feature_set = secondary_feature_set
    if len(in_feature_set) > 0:
        feature_set = in_feature_set

    if 'ans' not in in_data.columns and 'label' in in_data.columns:
        label = in_data['label'].to_list()
        in_data.insert(in_data.columns.get_loc("label") + 1, "ans", label)
        safe_drop(in_data, 'label')
    if 'prod' in in_data.columns:
        in_data['pred'] = in_data['prod']
        safe_drop(in_data, 'prod')

    if not is_typical:
        # Rename the features. If any of them is changed, overwrite the csv file.
        in_data, to_save = rename_features(in_data)
        if to_save:
            in_data.to_csv(path, index=False)
        renamed_feature_set = rename_feature_set(feature_set)
        # print(renamed_feature_set)

        # Filter out cases with known parameterization failures
        if to_filter:
            in_data = filter_df(in_data, FAILURES_YML)
    else:
        renamed_feature_set = feature_set

    in_data.replace([np.inf, -np.inf], np.nan, inplace=True)
    if in_data[in_data.isnull().any(axis=1)].shape[0] > 0:
        debug_nan_csv = './nan.csv'
        nan_df = in_data[in_data.isnull().any(axis=1)].copy()
        if os.path.exists(debug_nan_csv):
            df = pd.read_csv(path)
            nan_df = pd.concat([df, nan_df], ignore_index=True)
        nan_df.to_csv(debug_nan_csv, index=False)
        del nan_df

    data = pd.DataFrame()
    # Pick features
    if 'pred' in in_data.columns:
        data['pred'] = in_data['pred']
    data['stroke_indices1'] = in_data['stroke_indices1']
    data['stroke_indices2'] = in_data['stroke_indices2']
    data['ans'] = in_data['ans']

    # Order based on renamed_feature_set
    in_fea = []
    for col in in_data.columns:
        for fi, fea in enumerate(renamed_feature_set):
            if fea == re.sub(r'_b[0-9]*$', '_b', col):
                in_fea.append((fi, col))
                break
    in_fea = sorted(in_fea)
    in_fea = [f[1] for f in in_fea]
    for fea in in_fea:
        data[fea] = in_data[fea]
    # print(data.columns)

    # Drop infinity (implying no overlapping)
    data.dropna(how='any', inplace=True)
    # data.replace(np.inf, 0.0, inplace=True)

    full_data = in_data.loc[data.index, :]
    return data, full_data


def true_negative_rate(y_true, y_pred):
    tn, fp, _fn, _tp = confusion_matrix(y_true, y_pred, labels=[0, 1]).ravel()
    if (tn + fp) != 0:
        tnr = tn / (tn + fp)
    else:
        tnr = 1
    return tnr


def true_negative_rate_number(tn, fp):
    if (tn + fp) != 0:
        tnr = tn / (tn + fp)
    else:
        tnr = 1
    return tnr


def print_stats_from_confusion_mat(tn, fp, fn, tp, num_train, stage):
    C = np.array([[tn, fp],
                  [fn, tp]])
    table_list = [['Sample count',
                   tn + fp + fn + tp],
                  ['Failure count',
                   fp + fn],
                  ['| True positives',
                   tp],
                  ['| False positives',
                   fp],
                  ['| True negatives',
                   tn],
                  ['| False negatives',
                   fn],
                  ['Balanced accuracy',
                   balanced_accuracy_score_matrix(C)],
                  ['True negative rate',
                   true_negative_rate_number(tn, fp)],
                  ['F1 score',
                   f1_score_number(tp, fp, fn)],
                  ['Precision',
                   precision_score_number(tp, fp)],
                  ['Recall',
                   recall_score_number(tp, fn)]]
    if len(num_train) == 2:
        table_list = [
            ['Training count', num_train[0] + num_train[1]],
            ['| Positive training',
             num_train[0]],
            ['| Negative training',
             num_train[1]]
        ] + table_list
    tabu = tabulate(table_list,
                    headers=[f'({stage})', 'measure'],
                    floatfmt='.4f')
    print(tabu)
    return table_list


def print_stats(stage: str, GT: np.ndarray, Y: np.ndarray, num_train=[]):
    tn, fp, fn, tp = confusion_matrix(GT, Y, labels=[0, 1]).ravel()
    table_list = [['Sample count',
                   tn + fp + fn + tp],
                  ['Failure count',
                   fp + fn],
                  ['| True positives',
                   tp],
                  ['| False positives',
                   fp],
                  ['| True negatives',
                   tn],
                  ['| False negatives',
                   fn],
                  ['Balanced accuracy',
                   balanced_accuracy_score(GT, Y)],
                  ['True negative rate',
                   true_negative_rate(GT, Y)],
                  ['F1 score',
                   f1_score(GT, Y)],
                  ['Precision',
                   precision_score(GT, Y)],
                  ['Recall',
                   recall_score(GT, Y)]]
    if len(num_train) == 2:
        table_list = [
            ['Training count', num_train[0] + num_train[1]],
            ['| Positive training',
             num_train[0]],
            ['| Negative training',
             num_train[1]]
        ] + table_list
    tabu = tabulate(table_list,
                    headers=[f'({stage})', 'measure'],
                    floatfmt='.4f')
    print(tabu)
    return table_list


def balanced_accuracy_score_matrix(C):
    with np.errstate(divide="ignore", invalid="ignore"):
        per_class = np.diag(C) / C.sum(axis=1)
    if np.any(np.isnan(per_class)):
        per_class = per_class[~np.isnan(per_class)]
    score = np.mean(per_class)
    return score


def precision_score_number(tp, fp):
    if (tp + fp) == 0:
        return 1
    return tp / (tp + fp)


def recall_score_number(tp, fn):
    if (tp + fn) == 0:
        return 1
    return tp / (tp + fn)


def f1_score_number(tp, fp, fn):
    if (tp + 0.5 * (fp + fn)) == 0:
        return 1
    return tp / (tp + 0.5 * (fp + fn))


def compute_stats_table_list(C):
    tn = C[0, 0]
    fp = C[0, 1]
    fn = C[1, 0]
    tp = C[1, 1]

    out_table_list = [['Sample count',
                       tn + fp + fn + tp],
                      ['Failure count',
                       fp + fn],
                      ['| True positives',
                       tp],
                      ['| False positives',
                       fp],
                      ['| True negatives',
                       tn],
                      ['| False negatives',
                       fn],
                      ['Balanced accuracy',
                       balanced_accuracy_score_matrix(C)],
                      ['True negative rate',
                       true_negative_rate_number(tn, fp)],
                      ['F1 score',
                       f1_score_number(tp, fp, fn)],
                      ['Precision',
                       precision_score_number(tp, fp)],
                      ['Recall',
                       recall_score_number(tp, fn)]]
    return out_table_list


def sum_stats(table_list):
    tp = 0
    fp = 0
    tn = 0
    fn = 0
    for table in table_list:
        tp_i = [r[1] for r in table if '| True positives' == r[0]][0]
        fp_i = [r[1] for r in table if '| False positives' == r[0]][0]
        tn_i = [r[1] for r in table if '| True negatives' == r[0]][0]
        fn_i = [r[1] for r in table if '| False negatives' == r[0]][0]
        tp += tp_i
        fp += fp_i
        tn += tn_i
        fn += fn_i
    C = np.array([[tn, fp],
                  [fn, tp]])

    return compute_stats_table_list(C)


def print_feature_importance(classifier, fea_descriptions, X: np.ndarray, Y: np.ndarray):
    permu_importance = permutation_importance(classifier, X, Y)
    columns = zip(fea_descriptions, classifier.feature_importances_,
                  permu_importance.importances_mean, permu_importance.importances_std)
    print(tabulate([[f, f'{imp:.3f}', f'{p_imp:.3f} ± {p_imp_std:.3f}']
                    for (f, imp, p_imp, p_imp_std) in columns],
          headers=['Feature', 'Gini importance', 'Permu importance']))


def print_forest_stats(classifier):
    classifier_depths = [est.get_depth() for est in classifier.estimators_]
    classifier_leaves = [est.get_n_leaves() for est in classifier.estimators_]
    out_table = [
        ['# of trees', len(classifier.estimators_)],
        ['mean tree depth', statistics.mean(classifier_depths)],
        ['min tree depth',  min(classifier_depths)],
        ['max tree depth',  max(classifier_depths)],
        ['mean leaves',     statistics.mean(classifier_leaves)],
        ['min leaves',      min(classifier_leaves)],
        ['max leaves',      max(classifier_leaves)],
    ]
    print(tabulate(out_table, headers=['Name', 'Values']))

    return out_table


def determine_type(stroke_indices1, stroke_indices2):
    sindices1 = [int(s) for s in str(stroke_indices1).split(' ')]
    sindices2 = [int(s) for s in str(stroke_indices2).split(' ')]

    if len(sindices2) == 1:
        sindices1, sindices2 = sindices2, sindices1

    # Type 1: Single-single
    if len(sindices1) == 1 and len(sindices2) == 1:
        return 1
    elif len(sindices1) == 1 and len(sindices2) > 1:
        # Type 2: Single-multiple, single later than multiple
        if sindices1[0] > max(sindices2):
            return 2

        # Type 3: Single-multiple, single not later than multiple
        return 3
    elif len(sindices1) > 1 and len(sindices2) > 1:
        # Type 4: Multiple-multiple, two are clearly divided timewise
        if min(sindices1) > max(sindices2) or min(sindices2) > max(sindices1):
            return 4
        # Type 5: Multiple-multiple, two are not clearly divided timewise
        return 5
    else:
        raise Exception('Unseen type')


def add_columns(df, GT, prob):
    def safe_drop(data, col_name):
        try:
            data.drop(col_name, axis=1, inplace=True)
        except KeyError:
            print(f'Missing column: {col_name}')

    if "parameterization_success" in df.columns:
        df.insert(df.columns.get_loc(
            "parameterization_success") + 1, "label", GT)
    else:
        safe_drop(df, 'ans')
        if 'stroke_indices2' in df.columns:
            df.insert(df.columns.get_loc("stroke_indices2") + 1, "label", GT)
        else:
            df.insert(df.columns.get_loc("stroke_indices1") + 1, "label", GT)
    if len(prob) > 0:
        Y = (prob >= cutoff).astype(int)
        assert df.shape[0] == Y.shape[0]
        if 'type' not in df.columns and 'stroke_indices2' in df.columns:
            df.insert(df.columns.get_loc("label"), "type",
                      df.apply(lambda x: determine_type(x['stroke_indices1'], x['stroke_indices2']), axis=1))
        df.insert(df.columns.get_loc("label") + 1, "pred", Y)
        df.insert(df.columns.get_loc("pred") + 1, "prob", prob)
        df.insert(df.columns.get_loc("prob") + 1,
                  "success", GT.reshape(Y.shape) == Y)
    else:
        df.insert(df.columns.get_loc("label") + 1, "success", True)
    file_basename = [os.path.basename(
        f.replace('/input.svg', '')) for f in df['input']]
    df.insert(df.columns.get_loc("cluster_idx1"), "file", file_basename)

    if 'type' in df.columns:
        df = df.sort_values('type')

    return df


def read_measurement_examples(csv_file, is_pos, examples={}):
    example_pairs = pd.read_csv(csv_file, header=None, names=['col0', 'col1'])
    example_pairs_parsed = []
    for _, row in example_pairs.iterrows():
        row_list = []
        for col in example_pairs.columns:
            p = row[col]
            if isinstance(p, str) and ' ' in p:
                p = p.split(' ')
            else:
                p = [p]
            p = [int(s) for s in p]
            row_list.append(p)
        assert len(row_list) == 2
        example_pairs_parsed.append((row_list[0], row_list[1]))
    examples.update({1 if is_pos else 0: example_pairs_parsed})

    return examples


def read_measurement_examples_from_features(csv_file, examples={}):
    example_pairs, _ = preprocess_measure(csv_file, to_filter=True)
    # print(example_pairs)
    # exit()
    for _, row in example_pairs.iterrows():
        row_list = []
        for col in ['stroke_indices1', 'stroke_indices2']:
            p = row[col]
            if isinstance(p, str) and ' ' in p:
                p = p.split(' ')
            else:
                p = [p]
            p = [int(s) for s in p]
            row_list.append(p)
        assert len(row_list) == 2
        ans = row['ans']
        if ans not in examples:
            examples[ans] = []
        examples[ans].append((row_list[0], row_list[1]))

    return examples


def measure_clustering_on_examples(capture, examples: Dict[int, List[Tuple]]):
    if len(examples) == 0:
        return [['No examples', 0]], []

    error_examples = []

    # 1. Build clustering mapping
    s2c = {s.stroke_ind: s.group_ind for s in capture}

    tp = 0
    fp = 0
    tn = 0
    fn = 0

    def get_cluster(s):
        if s in s2c:
            return s2c[s]
        print(f'Missing stroke: #{s}')
        next_c = max(s2c.values())
        next_c += 1
        s2c[s] = next_c
        return next_c

    # 2. Check on examples
    assert 1 in examples or 0 in examples
    for k, es in examples.items():
        if k not in [0, 1]:
            continue

        # Check if there's any missing stroke
        no_missing_examples = []
        for example in es:
            has_missing = False
            for s in example[0]:
                if s not in s2c:
                    print(f'Missing stroke: #{s}')
                    has_missing = True
            for s in example[1]:
                if s not in s2c:
                    print(f'Missing stroke: #{s}')
                    has_missing = True

            if has_missing:
                if to_ignore_missing:
                    if k == 0:
                        tn += 1
                    else:
                        tp += 1
                else:
                    if k == 0:
                        fp += 1
                        error_examples.append(example)
                    else:
                        fn += 1
                        error_examples.append(example)
            else:
                no_missing_examples.append(example)

        # Negative examples: the two parts must be in two or more disjoint clusters.
        if k == 0:
            for example in no_missing_examples:
                cluster1 = set()
                for s in example[0]:
                    cluster1.add(get_cluster(s))
                cluster2 = set()
                for s in example[1]:
                    cluster2.add(get_cluster(s))

                if len(cluster1.intersection(cluster2)) > 0:
                    fp += 1
                    error_examples.append(example)
                    # print(example)
                    # print(cluster1)
                    # print(cluster2)
                    # print('-----------------')
                else:
                    tn += 1
        else:  # Positive examples: the two parts must be in the same cluster.
            for example in no_missing_examples:
                cluster1 = set()
                for s in example[0]:
                    cluster1.add(get_cluster(s))
                cluster2 = set()
                for s in example[1]:
                    cluster2.add(get_cluster(s))

                if len(cluster1.intersection(cluster2)) > 0:
                    tp += 1
                else:
                    fn += 1
                    error_examples.append(example)
                    # print(example)
                    # print(cluster1)
                    # print(cluster2)
                    # print('-----------------')

    # 3. Generate the statistic table
    C = np.array([[tn, fp],
                  [fn, tp]])

    return compute_stats_table_list(C), error_examples


#########################################################################


def format_sheet(ws):
    no_border = st.Border(left=st.Side(border_style=None),
                          right=st.Side(border_style=None),
                          top=st.Side(border_style=None),
                          bottom=st.Side(border_style=None),
                          diagonal=st.Side(border_style=None),
                          outline=st.Side(border_style=None),
                          vertical=st.Side(border_style=None),
                          horizontal=st.Side(border_style=None))
    right_border = st.Border(right=st.Side(
        border_style='thin', color='FF000000'))
    left_border = st.Border(left=st.Side(
        border_style='thin', color='FF000000'))
    header_alignment = openpyxl.styles.Alignment(
        horizontal='center', vertical='bottom', wrap_text=True)

    # Align header
    for cell in ws[1]:
        cell.alignment = header_alignment
        cell.border = no_border

    # Get column names
    column_names = {}
    for (i, col) in enumerate(ws.iter_cols(1, ws.max_column)):
        column_names[col[0].value] = openpyxl.utils.get_column_letter(i + 1)

    # Add vertical separation lines
    line_col = ['parameterization_success', 'stroke_indices2', 'success']
    if 'stroke_indices2' not in column_names:
        line_col = ['parameterization_success', 'stroke_indices1', 'success']
    for col_name in line_col:
        if col_name not in column_names:
            continue
        for cell in ws[column_names[col_name]]:
            cell.border = right_border
    for col_name in ('experiment_parameterization_result', 'input'):
        if col_name not in column_names:
            continue
        for cell in ws[column_names[col_name]]:
            cell.border = left_border

    # Precision
    for col_name in ('prob',):
        for cell in ws[column_names[col_name]]:
            cell.number_format = '0.00'

    return ws


def save_prediction_xlsx(df, GT, prob, xlsx_path):
    df = add_columns(df, GT, prob)

    with pd.ExcelWriter(xlsx_path) as xls_writer:
        df.to_excel(xls_writer, index=False,
                    sheet_name='Predictions', na_rep='-')

    xls = openpyxl.load_workbook(xlsx_path)

    ws = xls.active

    # Format sheet
    ws = format_sheet(ws)

    del ws

    print(f'Results saved as: {xlsx_path}')
    xls.save(xlsx_path)


def highlight_xlsx(df, GT, prob, highlight_cells, xlsx_path):
    df = add_columns(df, GT, prob)

    df_set = [df.loc[(df['label'] == 1)].copy(),
              df.loc[(df['label'] == 0)].copy()]
    del df
    with pd.ExcelWriter(xlsx_path) as xls_writer:
        # df.to_excel(xls_writer, index=False,
        #             sheet_name='Predictions', na_rep='-')
        df_set[0].to_excel(xls_writer, index=False,
                           sheet_name='Positive', na_rep='-')
        df_set[1].to_excel(xls_writer, index=False,
                           sheet_name='Negative', na_rep='-')

    xls = openpyxl.load_workbook(xlsx_path)

    for dff, sheet_name in zip(df_set, ['Positive', 'Negative']):
        dff.reset_index(drop=True, inplace=True)
        ws = xls[sheet_name]

        # Basic formatting
        ws = format_sheet(ws)

        # Highlight cells
        for cell_isoline, col_name in highlight_cells:
            row_idx = dff.index[dff['isoline'] == cell_isoline].tolist()
            assert len(row_idx) <= 1
            if len(row_idx) == 0:
                continue
            # Go from 0 indexing to 1 indexing, skip the header
            row_idx = row_idx[0] + 2
            # Go from 0 indexing to 1 indexing
            col_idx = dff.columns.get_loc(col_name) + 1
            cell_title = ws.cell(row_idx, col_idx)
            cell_title.fill = openpyxl.styles.fills.PatternFill(
                start_color="FFC7CE", fill_type="solid")

        del ws

    print(f'Results saved as: {xlsx_path}')
    xls.save(xlsx_path)


def feature_diff_xlsx(df, GT, prob, xlsx_path):
    df = add_columns(df, GT, prob)

    # Group features
    diff_strs = ['^average_', '^avg_', '^median_', '^mean_']
    diff_str = '|'.join(diff_strs)
    columns = [(re.sub(f'({diff_str})', '', col), i)
               for i, col in enumerate(df.columns)]
    column_groups = {}
    for col, i in columns:
        if col not in column_groups:
            column_groups[col] = []
        column_groups[col].append(i)
    column_groups = {k: cols for k,
                     cols in column_groups.items() if len(cols) > 1}
    print(column_groups)

    # Insert diff
    diff_df = df.copy()
    diff_insert_pos = df.columns.get_loc("success") + 1
    last_diff_col = ''
    for col, indices in column_groups.items():
        c1 = indices[0]
        c2 = indices[1]
        last_diff_col = 'diff_' + col
        df.insert(diff_insert_pos, last_diff_col,
                  (diff_df.iloc[:, c1] - diff_df.iloc[:, c2]).abs())
        diff_insert_pos += 1

    with pd.ExcelWriter(xlsx_path) as xls_writer:
        df.to_excel(xls_writer, index=False,
                    sheet_name='Difference', na_rep='-')

    xls = openpyxl.load_workbook(xlsx_path)
    ws = xls.active

    # Basic formatting
    ws = format_sheet(ws)

    # Get column names
    column_names = {}
    for (i, col) in enumerate(ws.iter_cols(1, ws.max_column)):
        column_names[col[0].value] = openpyxl.utils.get_column_letter(i + 1)
    right_border = st.Border(right=st.Side(
        border_style='thin', color='FF000000'))
    for cell in ws[column_names[last_diff_col]]:
        cell.border = right_border
    del ws

    print(f'Results saved as: {xlsx_path}')
    xls.save(xlsx_path)
